from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_report(results, stats, etalon_list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"

    # ----------------------------
    # Заголовки
    # ----------------------------
    ws["A1"] = "Распознаные номера"
    ws["B1"] = "Лучшее совпадение"
    ws["C1"] = "Дистанция"
    ws["E1"] = "Эталон не участвовал в сравнении"

    ws["G1"] = "Статистика качества детекции"
    ws["G2"] = "Всего распознано"
    ws["G3"] = "Полностью распознаные"
    ws["G4"] = "Частично распознаные"
    ws["G5"] = "Неверно распознано"
    ws["G6"] = "Полностью распознаные (%)"
    ws["G7"] = "Полностью + частично распознаные (%)"

    ws["H2"] = stats["total"]
    ws["H3"] = stats["fully_matched"]
    ws["H4"] = stats["partially_matched"]
    ws["H5"] = stats["incorrect"]
    ws["H6"] = round(stats["accuracy_percent"], 2)
    ws["H7"] = round(stats["full_plus_part"], 2)

    # ----------------------------
    # Цвета
    # ----------------------------
    green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    used_etalon = set()
    row_index = 2

    for item in results:
        recognized = ", ".join(item["recognized"])
        best = ", ".join(item["best_match"]) if item["best_match"] else ""
        dist = item["distance"]

        ws.cell(row=row_index, column=1, value=recognized)
        ws.cell(row=row_index, column=2, value=best)
        ws.cell(row=row_index, column=3, value=dist)

        if item["best_match"]:
            used_etalon.add(tuple(item["best_match"]))

        # Цветовая маркировка
        if dist == 0:
            fill = green
        elif dist <= 3:
            fill = yellow
        else:
            fill = red

        for col in range(1, 4):
            ws.cell(row=row_index, column=col).fill = fill

        row_index += 1

    # ----------------------------
    # Эталон не участвовал
    # ----------------------------
    not_used_row = 2

    for row in etalon_list:
        if tuple(row) not in used_etalon:
            ws.cell(row=not_used_row, column=5, value=", ".join(row))
            not_used_row += 1

    # =========================
    # Вторая статистика (с G10)
    # =========================
    ws["G10"] = "Статистика детекции"
    ws["G11"] = "Всего записей в эталоне"
    ws["G12"] = "Пропущенные номера"
    ws["G13"] = "Число повторов"

    ws["H11"] = stats["total_etalon"]
    ws["H12"] = stats["not_used_count"]
    ws["H13"] = stats["duplicates_count"]

    # =========================
    # Форматирование
    # =========================
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    alignment = Alignment(horizontal="center", vertical="center")

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border
                cell.alignment = alignment

    # ----------------------------
    # Автоширина столбцов
    # ----------------------------
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        max_length = 0

        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2

    wb.save(output_path)